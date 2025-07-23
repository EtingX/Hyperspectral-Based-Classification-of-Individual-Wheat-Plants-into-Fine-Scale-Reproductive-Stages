from classification_HL import *
import os.path
from openpyxl import load_workbook
import pandas as pd
import itertools

def split_data(excel_path, output_folder, seed=2024, plant_id_ranges=[(31, 70), (71, 120)],
               num_test_ids_per_range=10):
    try:
        # 读取Excel文件
        data = pd.read_excel(excel_path)
    except FileNotFoundError:
        print("指定的Excel文件未找到。")
        return
    except Exception as e:
        print(f"发生错误：{e}")
        return

    if 'Plant ID' not in data.columns:
        print("Excel文件中缺少'Plant ID'列。")
        return

    # 设置随机种子
    np.random.seed(seed)

    selected_test_ids = []

    # 处理每个范围
    for r in plant_id_ranges:
        # 从每个范围中随机选择20个Plant ID作为测试集ID
        ids_in_range = data[(data['Plant ID'] >= r[0]) & (data['Plant ID'] <= r[1])]['Plant ID'].unique()
        selected_test_ids.extend(np.random.choice(ids_in_range, size=num_test_ids_per_range, replace=False))

    print(selected_test_ids)
    # 判断每行数据是否应该进入测试集
    data['Set'] = data['Plant ID'].apply(lambda x: 'Test' if x in selected_test_ids else 'Train')

    # 分离训练集和测试集
    train_data = data[data['Set'] == 'Train']
    test_data = data[data['Set'] == 'Test']

    # 确保输出目录存在
    os.makedirs(output_folder, exist_ok=True)

    # 保存到指定文件夹
    train_data.to_excel(f"{output_folder}/train_data.xlsx", index=False)
    test_data.to_excel(f"{output_folder}/test_data.xlsx", index=False)

    # 打印出选中的测试集ID
    print("Selected Test IDs:", selected_test_ids)


# # 使用函數
# split_data(r"I:\0757 hyperspectral\separate dataset\labelled_data_0757 - three.xlsx", r'I:\0757 hyperspectral\separate dataset')

def hc2hhsi(hc):
    """
    Transform an n-dimensional hypercube to hyper-hue, saturation and intensity.

    :param hc: hyperCube(rows x cols x dims) in floating data type of [0 1].
    :return: hh: hypHue (rows x cols x (dims-1)) in floating data type of [0 1].
             s: saturation (rows x cols ) in floating data type of [0 1]
             i: intensity in (rows x cols) in floating data type of [0 1]

    For academic users, please cite:
    Citation 1:
    Liu, H., Lee, S., & Chahl, J.(2017).Transformation of a high-dimensional color space for material classification.
    Journal of the Optical Society of America A, 34(4), 523 - 532, doi: 10.1364/josaa.34.000523.

    Citation 2:
    Liu, H., & Chah, J.S.(2018).A multispectral machine vision system for invertebrate detection on green leaves
    Computer and Elecronics in Agriculture, 150, 279 - 288, doi: https://doi.org/10.1016/j.compag.2018.05.002.

    version 1.1 (Aug 18, 2018)
    Author: Huajian Liu
    """

    import numpy as np

    ####################################################################################################################
    # Calculate the components c
    rows = hc.shape[0]
    cols = hc.shape[1]
    dims = hc.shape[2]

    c = np.zeros((rows, cols, dims-1))
    for i in range(dims - 1):
        nonZeroEle = dims - i # nonZeroEle is the number of non-zero elements of the base unit vector u1, u2, ...
        c[:, :, i] = (nonZeroEle - 1) ** 0.5 / nonZeroEle ** 0.5         * hc[:, :, i] \
                     - 1 / ((nonZeroEle - 1) ** 0.5 * nonZeroEle ** 0.5) * np.sum(hc[:, :, i+1:dims], axis=2)
    ####################################################################################################################

    # Normalise the norms of c to 1 to obtain hyper-hue hh.
    c_norm = np.sum(c ** 2, axis=2) ** 0.5
    c_norm = c_norm + (c_norm == 0) * 1e-10
    c_norm = np.tile(c_norm, (dims - 1, 1, 1))
    c_norm = np.moveaxis(c_norm, 0, -1)
    hh = c / c_norm # add 1e-10 to avoid zero denominators

    # Saturation
    s = hc.max(2) - hc.min(2)
    # s = np.amax(hc, axis=2) - np.amin(hc, axis=2) # The same as above

    # Intensity
    i = 1/dims * np.sum(hc, 2)

    return hh, s, i

import numpy as np
from matplotlib import pyplot as plt
from sklearn.decomposition import PCA
def snv(ref, flag_fig=False):
    """
    Calculate the standard normal variate (SNV) of spectral signatures
    :param ref: 1D or 2D ndarray
    :param flag_fig: True or False to show the result
    :return: snv values
    """

    if ref.shape.__len__() == 1:
        ref = ref.reshape((1, ref.shape[0]))

    # meand and std
    mean_ref = np.mean(ref, axis=1)
    std_ref = np.std(ref, axis=1)

    mean_ref = mean_ref.reshape((mean_ref.shape[0], 1), order='C')
    std_ref = std_ref.reshape((std_ref.shape[0], 1), order='C')

    mean_ref = np.tile(mean_ref, (1, ref.shape[1]))
    std_ref = np.tile(std_ref, (1, ref.shape[1]))

    # snv
    snv = (ref - mean_ref) / std_ref

    if flag_fig:
        f = plt.figure()
        a1 = f.add_subplot(1, 2, 1)
        a2 = f.add_subplot(1, 2, 2)

        for a_ref in ref:
            a1.plot(a_ref)
            a1.set_title('Reflectance')

        for a_snv in snv:
            a2.plot(a_snv)
            a2.set_title('SNV')

    return snv
def apply_pca(ref):
    pca = PCA(n_components=20)
    transformed_data = pca.fit_transform(ref)
    return transformed_data

train_output_path = r"I:\paper 2 analyse\ASD dataset\test_data_average_resample.xlsx"
# test_output_path = r"I:\ASD hyper code GAN\train_data_mid.xlsx"

# hyp, snv, pca in train
df_train = pd.read_excel(train_output_path)

# Load the workbook and rename the first sheet to 'org'
wb = load_workbook(train_output_path)
sheet_names = wb.sheetnames
wb[sheet_names[0]].title = 'org'
wb.save(train_output_path)
wb.close()

# Continue processing
# Select columns for transformation and preserve the first 4 and last 2 columns
ref_train = df_train.iloc[:, 3:-2].values

ref_train = ref_train.astype(float)

first_cols = df_train.iloc[:, :3]
last_cols = df_train.iloc[:, -2:]

# Perform transformations
pca_data = apply_pca(ref_train)

snv_data = snv(ref_train)

hc = ref_train.reshape(ref_train.shape[0], 1, ref_train.shape[1])
hh, sat, inten = hc2hhsi(hc)
hh_data = hh.reshape(hh.shape[0], hh.shape[2])

# Create DataFrames for the transformed data
pca_df = pd.DataFrame(pca_data, columns=[f'PCA_Feature_{i}' for i in range(pca_data.shape[1])])
snv_df = pd.DataFrame(snv_data, columns=[f'SNV_Feature_{i}' for i in range(snv_data.shape[1])])
hh_df = pd.DataFrame(hh_data, columns=[f'HH_Feature_{i}' for i in range(hh_data.shape[1])])
hh_df['Saturation'] = sat
hh_df['Intensity'] = inten

# Concatenate first 4 columns, transformed data, and last 2 columns in order
final_pca_df = pd.concat([first_cols, pca_df, last_cols], axis=1)
final_snv_df = pd.concat([first_cols, snv_df, last_cols], axis=1)
final_hh_df = pd.concat([first_cols, hh_df, last_cols], axis=1)

# Save to the existing Excel file in new sheets using openpyxl for compatibility
with pd.ExcelWriter(train_output_path, mode='a', engine='openpyxl', if_sheet_exists='replace') as writer:
    final_hh_df.to_excel(writer, sheet_name='hyp', index=False)
    final_snv_df.to_excel(writer, sheet_name='snv', index=False)
    final_pca_df.to_excel(writer, sheet_name='pca', index=False)
#
# # hyp, snv, pca in test
# df_test = pd.read_excel(test_output_path)
#
# # Load the workbook and rename the first sheet to 'org'
# wb = load_workbook(test_output_path)
# sheet_names = wb.sheetnames
# wb[sheet_names[0]].title = 'org'
# wb.save(test_output_path)
# wb.close()
#
# # Continue processing
# # Select columns for transformation and preserve the first 4 and last 2 columns
# ref_test = df_test.iloc[:, 4:-2].values
#
# ref_test = ref_test.astype(float)
#
# first_cols_test = df_test.iloc[:, :4]
# last_cols_test = df_test.iloc[:, -2:]
#
# # Perform transformations
# pca_data_test = apply_pca(ref_test)
#
# snv_data_test = snv(ref_test)
#
# hc_test = ref_test.reshape(ref_test.shape[0], 1, ref_test.shape[1])
# hh_test, sat_test, inten_test = hc2hhsi(hc_test)
# hh_data_test = hh_test.reshape(hh_test.shape[0], hh_test.shape[2])
#
# # Create DataFrames for the transformed data
# pca_df_test = pd.DataFrame(pca_data_test, columns=[f'PCA_Feature_{i}' for i in range(pca_data_test.shape[1])])
# snv_df_test = pd.DataFrame(snv_data_test, columns=[f'SNV_Feature_{i}' for i in range(snv_data_test.shape[1])])
# hh_df_test = pd.DataFrame(hh_data_test, columns=[f'HH_Feature_{i}' for i in range(hh_data_test.shape[1])])
# hh_df_test['Saturation'] = sat_test
# hh_df_test['Intensity'] = inten_test
#
# # Concatenate first 4 columns, transformed data, and last 2 columns in order
# final_pca_df_test = pd.concat([first_cols_test, pca_df_test, last_cols_test], axis=1)
# final_snv_df_test = pd.concat([first_cols_test, snv_df_test, last_cols_test], axis=1)
# final_hh_df_test = pd.concat([first_cols_test, hh_df_test, last_cols_test], axis=1)
#
# # Save to the existing Excel file in new sheets using openpyxl for compatibility
# with pd.ExcelWriter(test_output_path, mode='a', engine='openpyxl', if_sheet_exists='replace') as writer:
#     final_hh_df_test.to_excel(writer, sheet_name='hyp', index=False)
#     final_snv_df_test.to_excel(writer, sheet_name='snv', index=False)
#     final_pca_df_test.to_excel(writer, sheet_name='pca', index=False)
#
# print('----------------------------------------------------------------')
# print('Transform finished!')